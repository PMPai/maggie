from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from decimal import Decimal


def generate_billing_excel(
    output_path: str,
    company_name: str,
    owner_name: str,
    project_name: str,
    contract_no: str,
    application_date: str,
    period_no: int,
    contract_total: str,
    lines: list[dict],
    totals: dict,
) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = f"第{period_no}期请款单"

    # Header
    ws["A1"] = company_name
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells("A1:M1")
    ws["A2"] = "工程计价请款单"
    ws["A2"].font = Font(size=12, bold=True)
    ws.merge_cells("A2:M2")
    ws["A3"] = f"业主：{owner_name}    工程名称：{project_name}    合同编号：{contract_no}"
    ws.merge_cells("A3:M3")
    ws["A4"] = f"请款日期：{application_date}    请款期数：第{period_no}期    合同总价：{contract_total}"
    ws.merge_cells("A4:M4")

    headers = ["项次", "项目名称", "单位", "合同数量", "单价", "复价", "前期累计", "本期数量", "累计数量", "施工金额", "保留款", "计价金额", "备注"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="F0F0F0", fill_type="solid")

    row_num = 7
    for line in lines:
        ws.cell(row=row_num, column=1, value=line.get("line_no", ""))
        ws.cell(row=row_num, column=2, value=line.get("description", ""))
        ws.cell(row=row_num, column=3, value=line.get("unit", ""))
        ws.cell(row=row_num, column=4, value=float(line.get("contract_qty", 0)))
        ws.cell(row=row_num, column=5, value=float(line.get("unit_price", 0)))
        ws.cell(row=row_num, column=6, value=float(line.get("line_amount", 0)))
        ws.cell(row=row_num, column=7, value=float(line.get("prev_qty", 0)))
        ws.cell(row=row_num, column=8, value=float(line.get("current_qty", 0)))
        ws.cell(row=row_num, column=9, value=float(line.get("cumulative_qty", 0)))
        ws.cell(row=row_num, column=10, value=float(line.get("work_amount", 0)))
        ws.cell(row=row_num, column=11, value=float(line.get("retention", 0)))
        ws.cell(row=row_num, column=12, value=float(line.get("billed_amount", 0)))
        ws.cell(row=row_num, column=13, value=line.get("remarks", ""))
        for col in range(4, 13):
            ws.cell(row=row_num, column=col).number_format = "#,##0.00"
            ws.cell(row=row_num, column=col).alignment = Alignment(horizontal="right")
        row_num += 1

    # Totals
    row_num += 1
    total_labels = [("本期施工金额", totals["gross"]), ("本期保留款", totals["retention_held"]),
                    ("本期释放保留款", totals["retention_released"]), ("本期扣款", totals["deduction"]),
                    ("未税计价金额", totals["taxable"]), ("税额", totals["tax"]),
                    ("含税发票金额", totals["invoice"])]
    for label, val in total_labels:
        ws.cell(row=row_num, column=11, value=label).font = Font(bold=True)
        cell = ws.cell(row=row_num, column=12, value=float(val))
        cell.font = Font(bold=True)
        cell.number_format = "#,##0.00"
        cell.alignment = Alignment(horizontal="right")
        row_num += 1

    # Column widths
    for col, width in enumerate([8, 30, 8, 12, 12, 12, 12, 12, 12, 14, 12, 14, 20], 1):
        ws.column_dimensions[chr(64 + col) if col <= 26 else "A"].width = width

    wb.save(output_path)
    return output_path
